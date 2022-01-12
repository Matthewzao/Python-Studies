from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select

import time
import openpyxl
from openpyxl import load_workbook

chromeDriver = "C:\\Users\\Up Agency 2\\OneDrive\\Área de Trabalho\\estudo python\\selenium\\chromedriver.exe"
driver = webdriver.Chrome(chromeDriver)
driver.maximize_window()

wb = load_workbook('C:\\Users\\Up Agency 2\\OneDrive\\Área de Trabalho\\estudo python\\selenium\\SUBIR PREÇO LE BRANDS TEST.xlsx')
sheet = wb.active
rows = sheet.max_row


driver.get('https://www.lboficial.com.br/admin')
time.sleep(0.5)

driver.find_element_by_xpath('//*[@id="UsuarioEmail"]').send_keys('suporte@upagencybrasil.com.br')
driver.find_element_by_xpath('//*[@id="UsuarioSenha"]').send_keys('@VoaUpagency2021!')
driver.find_element_by_xpath('//*[@id="frm-login"]/div[4]/input').click()
time.sleep(0.5)

for i in range(1, rows+1): 
  ref = sheet.cell(row=i, column=2).value
    preco = sheet.cell(row=i, column=9).value
    driver.get('https://www.lboficial.com.br/admin/produtos')  
    time.sleep(0.5)
    driver.find_element_by_xpath('//*[@id="artigo"]/section[1]/ul/li[2]/a').click()
    driver.find_element_by_xpath('//*[@id="artigo"]/section[1]/ul/li[2]/div/ul/li[3]/label').click()
    driver.find_element_by_xpath('//*[@id="Produto-referencia"]').send_keys(ref)
    driver.find_element_by_xpath('//*[@id="form-filtro"]/input').click()
    driver.find_element_by_xpath('//*[@id="artigo"]/section[5]/div/table/tbody/tr/td[4]/div/div[1]').click()
    driver.find_element_by_xpath('//*[@id="artigo"]/section[5]/div/table/tbody/tr[1]/td[4]/div/div[2]/a[1]').click()
    time.sleep(0.5)

    driver.find_element_by_xpath('//*[@id="form"]/div/div[2]/div[1]/section[1]/ul/li/a').click()
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="form"]/fieldset/div/ul[1]/li').click()
    driver.find_element_by_xpath('//*[@id="btn-gravar"]').click()
    time.sleep(1)