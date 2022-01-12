from openpyxl.workbook import workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select

import time
import openpyxl
from openpyxl import load_workbook

chromeDriver = ('C:\\Users\\Up Agency 2\\OneDrive\\Área de Trabalho\\estudo python\\selenium.exe')
driver = webdriver.Chrome(chromeDriver)
driver.maximize_window()

wb = load_workbook('C:\Users\Up Agency 2\OneDrive\Área de Trabalho\\apagar\BLACKPINK - Relatório Geral_PRODUTOS_Tabela.xlsx')
sheet = wb.active
rows = sheet.max_row
columns = sheet.max_column

driver.get("https://idrissidress.braavo.com.br/admin")
username = driver.find_element_by_xpath('//*[@id="UsuarioEmail"]')
username.send_keys("suporte@upagencybrasil.com.br")
password = driver.find_element_by_xpath('//*[@id="UsuarioSenha"]')
password.send_keys("@VoaUpagency2021!")
driver.find_element_by_xpath('//*[@id="frm-login"]/div[4]/input').click()

for i in range(2, rows+1):
    codigo =  sheet.cell(row=i, column=2).value
    driver.get('https://www.blackpinkshop.com.br/admin/produtos/editar/{}'.format(codigo))
    driver.find_element_by_xpath('//*[@id="ProdutoAtivo2"]').click()
    driver.find_element_by_xpath('//*[@id="btn-gravar"]').click()