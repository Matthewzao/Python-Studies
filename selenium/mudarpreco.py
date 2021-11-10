from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select

import time
import openpyxl
from openpyxl import load_workbook

chromeDriver = "C:\\Users\\Up Agency 2\\OneDrive\\Área de Trabalho\\estudo python\\selenium\\chromedriver.exe"
driver = webdriver.Chrome(chromeDriver)
driver.maximize_window()

wb = load_workbook('C:\\Users\\Up Agency 2\\OneDrive\\Área de Trabalho\\estudo python\\selenium\\produtosBlackPink.xlsx')
sheet = wb.active
rows = sheet.max_row


driver.get('https://www.blackpinkshop.com.br/admin')
time.sleep(0.5)

driver.find_element_by_xpath('//*[@id="UsuarioEmail"]').send_keys('suporte@upagencybrasil.com.br')
driver.find_element_by_xpath('//*[@id="UsuarioSenha"]').send_keys('@VoaUpagency2021!')
driver.find_element_by_xpath('//*[@id="frm-login"]/div[4]/input').click()
time.sleep(0.5)

for i in range(2, rows+1): 
    codigo = sheet.cell(row=i, column=1).value
    preco = sheet.cell(row=i, column=2).value
    driver.get('https://www.blackpinkshop.com.br/admin/skus/editar/{}'.format(codigo))
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="SkuPrecoVenda"]').clear()
    driver.find_element_by_xpath('//*[@id="SkuPrecoVenda"]').send_keys(preco)
    driver.find_element_by_xpath('//*[@id="SkuPrecoVenda"]').send_keys("00")

    driver.find_element_by_xpath('//*[@id="btn-gravar"]').click()

    time.sleep(1)