from typing import Collection
from openpyxl.cell import cell
from openpyxl.styles.styleable import NamedStyleDescriptor
import requests
import json
import time
from requests import auth
from requests.api import request
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl import load_workbook

# tratamento planilha
wb = load_workbook('C:\\Users\\Up Agency 2\\OneDrive\\Área de Trabalho\\desconto\\desconto teste.xlsx')
sheet = wb.active
rows = sheet.max_row
columns = sheet.max_column

#entrando na API

token = '6a37d792f83103219113144428d0b19a'  
key = 'e4594fbffca3b6a8a0efd23de753b58a' 
url_site = "https://www.estiloplushe.com.br/api/produtos_descontos/editar" 


for r in range(2, rows+1):
    id = sheet.cell(row=r, column=1).value
    desc = sheet.cell(row=r, column=2).value
    desconto = {
    "produto_id": id,
    "desconto_tipo": "P",
    "desconto_valor": 0,
    "data_desconto_1": "2010-01-01",
    "data_desconto_2": "2050-12-31"
    }
    addDesconto = requests.post(url=url_site, auth=HTTPBasicAuth(token,key),json=desconto)
    print(addDesconto.json())
    print.("produto apagado")