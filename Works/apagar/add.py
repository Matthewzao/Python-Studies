from typing import Collection
from openpyxl.cell import cell
from openpyxl.styles.styleable import NamedStyleDescriptor
import requests
import json
import time
from requests import auth
from requests.api import request
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl import load_workbook

# tratamento planilha
wb = load_workbook('C:\\Users\\gu\\Desktop\\CriaCor\\estoque millennium 081221.xlsx')
sheet = wb.active
rows = sheet.max_row
columns = sheet.max_column

# entrando na API
token = '90f3237005640b0688e4f5fb3d8e9c52'
key = '388ecbdcc8e26e5dff55390b87f75be3'

#__ADICIONA COR_#

for i in range(70, rows+1):
    nomeCor = sheet.cell(row=i, column=1).value
    cor = {
    "pai_id": "1001",
    "ativo": "1",
    "nome": nomeCor
    }
    addCor = requests.post(url="https://idrissidress.braavo.com.br/api/variacoes/adicionar", auth=HTTPBasicAuth(token,key), json=cor)
    print(addCor.json())


#__EXCLUI COR__#

for i in range(1, rows+1):
    idCor = sheet.cell(row=i, column=1).value
    cor = {
        'id':idCor
    }
    removeCor = requests.post(url="https://idrissidress.braavo.com.br/api/variacoes/apagar",auth=HTTPBasicAuth(token,key),json=cor)
    print(removeCor.json())


#___EXCLUI SKU__#

for n in range(858,943):
    sku = {
        'id':n
    }
    removeSku = requests.post(url="https://idrissidress.braavo.com.br/api/skus/apagar",auth=HTTPBasicAuth(token,key),json=sku)
    print(removeSku.json())