from typing import Collection
from openpyxl.cell import cell
from openpyxl.styles.styleable import NamedStyleDescriptor
import requests
import json
import time
from requests import auth
from requests.api import request
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl import load_workbook

# tratamento planilha
wb = load_workbook('C:\\Users\\Up Agency 2\\OneDrive\\Área de Trabalho\\SUBIR RUSH\\RUSH_RUSH_1.xlsx')
sheet = wb.active
rows = sheet.max_row
columns = sheet.max_column

#entrando na API
token = '4423f2d8922cad3e3ec5a8c342992c04'   
key = '1582c527884345cd8153fda9b11c7356' 
url_site = "https://www.rushrush.com.br/api/produtos/adicionar" 


names = []
refs = []
ncms = []
descs = []
precos = []
cats = []

for r in range(2, rows+1):
    if(sheet.cell(row=r, column=2).value != sheet.cell(row=r+1, column=2).value):
        names.append(sheet.cell(row=r, column=1).value)
    if(sheet.cell(row=r, column=2).value != sheet.cell(row=r+1, column=2).value):
        refs.append(sheet.cell(row=r, column=2).value)
    if(sheet.cell(row=r, column=2).value != sheet.cell(row=r+1, column=2).value):
        ncms.append(sheet.cell(row=r, column=5).value) 
    if(sheet.cell(row=r, column=2).value != sheet.cell(row=r+1, column=2).value):
        cats.append(sheet.cell(row=r, column=6).value)  
    if(sheet.cell(row=r, column=2).value != sheet.cell(row=r+1, column=2).value):
        precos.append(sheet.cell(row=r, column=11).value)   
    if(sheet.cell(row=r, column=2).value != sheet.cell(row=r+1, column=2).value):
        descs.append(sheet.cell(row=r, column=12).value)  

for i in range(0, len(names)):
    produto = {
        "categoria_id": cats[i],
        "ativo": "1",
        "nome": names[i],
        "ref": refs[i], 
        "ncm": ncms[i],
        "deduzir": "1",
        "esgotado": "1",
        "prazo": "0",
        "descricao": descs[i]
    }   
    print(names[i])
    addProduto = requests.post(url=url_site, auth=HTTPBasicAuth(token,key),json=produto)
    time.sleep(1)
    addProdutoJson = addProduto.json()
    print(addProdutoJson)
    idProduto = addProdutoJson['ok']["id"]

    for k in range(2, rows+1):
        if(sheet.cell(row=k, column=2).value == refs[i]):
            valueSku = sheet.cell(row=k, column=3).value
            var1 = sheet.cell(row=k, column=4).value
            var2 = sheet.cell(row=k,column=7).value
            peso = sheet.cell(row=k,column=9).value
            qtd = sheet.cell(row=k, column=8).value
            sku = {
                "produto_id": idProduto,
                "var1_id": var1,
                "var2_id": var2,
                "ativo": "1",
                "ref": valueSku,
                "peso": peso,
                "preco_custo": "00.00",
                "preco_venda": precos[i],
                "movimentacao": "entrada",
                "lancamento": qtd,
                "gtin": "",
                "mpn": ""
            }
            addSku = requests.post(url='https://www.rushrush.com.br/api/skus/adicionar', auth=HTTPBasicAuth(token,key), json=sku)

            addSkuJson = addSku.json()
            time.sleep(1)
            print(addSkuJson)
            print('sku adicionado')
    print("produto add")
print("finished")